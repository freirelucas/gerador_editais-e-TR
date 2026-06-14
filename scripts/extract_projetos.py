#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Processa a planilha de projetos do IPEA (Projetos ativos) e gera src/data/projetos.js
— seeds reais para o Gerador de TR/Edital.

Cada projeto ATIVO (não Concluído/Cancelado) vira um conjunto de parâmetros do wizard,
sempre dentro do UNIVERSO DA BASE (taxonomia de tema/função e as 6 diretorias substantivas
de pesquisa). Quando o projeto não tem precedente na base — diretoria não-substantiva
(DIDES, GABIN, CGDTI…) ou título que não casa com nenhum tema —, COMPOMOS a partir do
próprio projeto (diretorias envolvidas, tema da área) e marcamos o campo como composto.

Uso:  python scripts/extract_projetos.py  CAMINHO_DO_XLSX
(sem argumento, usa o upload padrão desta sessão.)
"""
import sys, json, re, unicodedata
from pathlib import Path
from collections import Counter, defaultdict

RAIZ = Path(__file__).resolve().parent.parent
XLSX = sys.argv[1] if len(sys.argv) > 1 else \
    "/root/.claude/uploads/a857d93b-a2a0-514d-a37b-f734238a4c84/ff466111-14062026_projetos_1.xlsx"
TAX = json.loads((RAIZ / "data/taxonomia.json").read_text(encoding="utf-8"))

# 6 diretorias substantivas de pesquisa (universo da base) e o tema da área.
DIRETORIA_TEMA = {
    "DIEST": "Estado, instituições e democracia",
    "DIMAC": "Macroeconomia e finanças",
    "DISOC": "Social, trabalho e renda",
    "DIRUR": "Regional, urbano e ambiental",
    "DISET": "Setorial, inovação e infraestrutura",
    "DINTE": "Internacional e comércio",
}
SUBSTANTIVAS = set(DIRETORIA_TEMA)
TEMA_DIRETORIA = {t: s for s, t in DIRETORIA_TEMA.items()}
ATIVO_FORA = ("conclu", "cancel")  # estágios que NÃO são ativos


def sa(s):  # sem acento, minúsculo (mesma convenção da taxonomia)
    s = unicodedata.normalize("NFD", str(s or ""))
    return "".join(c for c in s if unicodedata.category(c) != "Mn").lower()


def classifica(texto, grupo):
    """Multi-rótulo por substring (sem acento), na ordem da taxonomia."""
    t = sa(texto)
    return [cat for cat, kws in TAX[grupo].items() if any(sa(k) in t for k in kws)]


def carrega_linhas():
    from openpyxl import load_workbook
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb["Projetos"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(c) for c in rows[0]]
    idx = {h: i for i, h in enumerate(hdr)}

    def col(row, *names):
        for n in names:
            i = next((idx[h] for h in hdr if n.lower() in h.lower()), None)
            if i is not None and i < len(row) and row[i] not in (None, ""):
                return row[i]
        return None
    return rows[1:], col


def sigla_substantiva(coord, envolvidas, temas):
    """Diretoria dentro do universo da base. Coordenadora se substantiva; senão a 1ª
    substantiva entre as envolvidas; senão a do tema classificado; senão DIEST (padrão)."""
    composto = False
    c = (coord or "").strip().upper()
    if c in SUBSTANTIVAS:
        return c, composto
    composto = True
    for s in re.split(r"[;,/]", (envolvidas or "")):
        s = s.strip().upper()
        if s in SUBSTANTIVAS:
            return s, composto
    if temas:
        s = TEMA_DIRETORIA.get(temas[0])
        if s:
            return s, composto
    return "DIEST", composto


def ler_produtos():
    """Produtos por 'Número do Projeto' (Counter de tipo) — sheet Produtos."""
    from openpyxl import load_workbook
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    rows = wb["Produtos"].iter_rows(values_only=True)
    hdr = [str(c) for c in next(rows)]
    inum = next(i for i, h in enumerate(hdr) if "número do projeto" in h.lower())
    itipo = next(i for i, h in enumerate(hdr) if "tipo de produto" in h.lower())
    prod = {}
    for r in rows:
        if r[itipo]:
            prod.setdefault(str(r[inum]), Counter())[r[itipo]] += 1
    return prod


# Inferência de FUNÇÃO pelo MIX DE PRODUTOS quando o título não a revela (≈90% dos casos).
# O único sinal específico e defensável dos produtos é "Base de Dados" → engenharia de
# dados. As demais saídas (artigos, TDs, notas, docs institucionais) são apoio à pesquisa
# (revisão, sistematização, relatórios — FUNCAO_TERMOS de "Apoio à pesquisa"). A distinção
# fina quant/DS vem do título. Sem produtos → função fica "a definir".
PROD_DATA = "Base de Dados"
def funcao_por_produtos(cnt):
    if not cnt:
        return []
    tot = sum(cnt.values())
    if cnt.get(PROD_DATA, 0) >= 2 and cnt[PROD_DATA] / tot >= 0.15:
        return ["Ciência de Dados – Engenharia"]
    return ["Apoio à pesquisa"]


def main():
    linhas, col = carrega_linhas()
    prod = ler_produtos()  # produtos por Número do Projeto, p/ inferir função e o resumo
    seeds, vistos, n_estrat, n_prior = [], set(), 0, 0
    for r in linhas:
        estagio = col(r, "Estágio do Projeto")
        if not estagio or any(x in sa(estagio) for x in ATIVO_FORA):
            continue  # só projetos ativos
        titulo = (col(r, "Título") or "").strip()
        numero = (col(r, "Número do Projeto") or "").strip()
        if not titulo or numero in vistos:
            continue
        vistos.add(numero)
        if sa(col(r, "Projeto Estratégico") or "") == "estrategico": n_estrat += 1
        if sa(col(r, "Projeto Prioritário") or "") == "prioritario": n_prior += 1
        coord_dir = col(r, "Diretoria")
        envolvidas = col(r, "Diretorias Envolvidas")
        # classificação dentro do universo da base (título do projeto)
        temas = classifica(titulo, "TEMA")
        titulo_func = classifica(titulo, "FUNCAO")
        if titulo_func:                       # título revela a função (alta confiança)
            funcoes, func_composto = titulo_func, False
        else:                                 # senão, infere pelo mix de produtos
            funcoes = funcao_por_produtos(prod.get(numero))
            func_composto = bool(funcoes)
        diretoria, dir_composto = sigla_substantiva(coord_dir, envolvidas, temas)
        tema_composto = not temas
        if not temas:  # sem precedente no título → compõe pelo tema da área
            temas = [DIRETORIA_TEMA[diretoria]]
        data_ini = col(r, "Data Inicial")
        ano = str(data_ini)[:4] if data_ini else None
        seeds.append({
            "id": str(col(r, "ID Projeto") or numero),
            "numero": numero,
            "titulo": titulo,
            "coordenador": (col(r, "Coordenador do projeto") or "").strip() or None,
            "diretoria": diretoria,
            "diretoriaRaw": (coord_dir or "").strip() or None,
            "temas": temas,
            "funcoes": funcoes,
            "modalidadeProjeto": (col(r, "Modalidade") or "").strip() or None,
            "estagio": estagio,
            "ano": ano,
            "composto": {"diretoria": dir_composto, "tema": tema_composto, "funcao": func_composto},
        })

    # ordena por diretoria → tema → ano desc, p/ navegação previsível no picker
    seeds.sort(key=lambda s: (s["diretoria"], s["temas"][0], s["ano"] or "0"), reverse=False)

    prod_ativos = Counter()
    for n in vistos:
        prod_ativos.update(prod.get(n, Counter()))
    tot_prod, prod_tipos = sum(prod_ativos.values()), prod_ativos.most_common(12)
    # cobertura de bolsas por tema (sinal de oportunidade): chamadas PIPA ÷ projetos ativos
    corpus = json.loads((RAIZ / "data/corpus_chamadas_2023-2026.json").read_text(encoding="utf-8"))
    pipa_tema = Counter(t for c in corpus if c.get("programa") == "PIPA" for t in (c.get("categoria_tema") or []))
    proj_tema = Counter(s["temas"][0] for s in seeds)
    cobertura = {t: {"ativos": proj_tema[t], "pipa": pipa_tema.get(t, 0)} for t in proj_tema}
    resumo = {
        "totalAtivos": len(seeds),
        "totalProdutos": tot_prod,
        "produtosPorTipo": [{"label": t, "value": n} for t, n in prod_tipos],
        "coberturaPorTema": cobertura,
        "estrategicos": n_estrat,
        "prioritarios": n_prior,
    }
    out = RAIZ / "src/data/projetos.js"
    head = (
        "// GERADO por scripts/extract_projetos.py — projetos ATIVOS do IPEA como seeds do Gerador.\n"
        "// Cada projeto traz parâmetros do wizard no UNIVERSO DA BASE (taxonomia de tema/função +\n"
        "// 6 diretorias substantivas). `composto.diretoria/tema = true` quando não havia precedente\n"
        "// e o valor foi composto a partir do próprio projeto: diretorias envolvidas / tema da\n"
        "// área / função inferida pelo mix de produtos (Base de Dados → engenharia; demais saídas\n"
        "// → apoio à pesquisa). PROJETOS_RESUMO agrega os produtos entregues (sheet Produtos) p/ o\n"
        "// Analytics. Não edite à mão; rode o script.\n"
        f"export const PROJETOS = {json.dumps(seeds, ensure_ascii=False, indent=0)};\n"
        f"export const PROJETOS_RESUMO = {json.dumps(resumo, ensure_ascii=False, indent=0)};\n"
    )
    out.write_text(head, encoding="utf-8")

    # ---- relatório de cobertura (diversidade das combinações) ----
    combos = Counter((s["diretoria"], s["temas"][0], (s["funcoes"][0] if s["funcoes"] else "—")) for s in seeds)
    print(f"projetos ativos: {len(seeds)}  →  {out.relative_to(RAIZ)}")
    print("por diretoria:", dict(Counter(s["diretoria"] for s in seeds)))
    print("por tema(1º):", dict(Counter(s["temas"][0] for s in seeds)))
    print("com função (total):", sum(1 for s in seeds if s["funcoes"]),
          "| pelo título:", sum(1 for s in seeds if s["funcoes"] and not s["composto"]["funcao"]),
          "| inferida de produtos:", sum(1 for s in seeds if s["composto"]["funcao"]),
          "| a definir:", sum(1 for s in seeds if not s["funcoes"]))
    print("função (distribuição):", dict(Counter(s["funcoes"][0] for s in seeds if s["funcoes"]).most_common()))
    print("diretoria composta:", sum(1 for s in seeds if s["composto"]["diretoria"]),
          "| tema composto:", sum(1 for s in seeds if s["composto"]["tema"]))
    print(f"combinações distintas (diretoria×tema×função): {len(combos)}")


if __name__ == "__main__":
    main()
